from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from dbhelper import DBHelper
from datetime import datetime
import re
import io
import openpyxl
from openpyxl.utils import get_column_letter
from helpers import extract_chapters

hojas_trabajo_bp = Blueprint("hojas_trabajo_bp", __name__)
db = DBHelper()

@hojas_trabajo_bp.route("/hojas_trabajo")
def listar_hojas():
    hojas = db.execute_query("""
        SELECT h.*, pr.nombre_proyecto, cl.nombre AS cliente,
            COALESCE(h.tecnico_encargado, 
                (SELECT tecnico_encargado FROM presupuestos WHERE id_proyecto = h.id_proyecto ORDER BY id DESC LIMIT 1)
            ) AS tecnico_encargado
        FROM hojas_trabajo h
        JOIN proyectos pr ON h.id_proyecto = pr.id
        JOIN clientes cl ON pr.id_cliente = cl.id
        ORDER BY h.id DESC
    """, fetch=True)
    return render_template("hojas_trabajo_list.html", hojas=hojas)

@hojas_trabajo_bp.route("/hojas_trabajo/nuevo", methods=["GET", "POST"])
def nuevo_hoja_trabajo():
    flash("Funcionalidad de 'Nueva Hoja de Trabajo' aún no implementada.", "info")
    return redirect(url_for("hojas_trabajo_bp.listar_hojas"))

@hojas_trabajo_bp.route("/hojas_trabajo/editar/<int:id>", methods=["GET", "POST"])
def editar_hoja_trabajo(id):
    hoja = db.execute_query("""
        SELECT h.*, pr.nombre_proyecto, cl.nombre AS cliente,
            COALESCE(h.tecnico_encargado, 
                (SELECT tecnico_encargado FROM presupuestos WHERE id_proyecto = h.id_proyecto ORDER BY id DESC LIMIT 1)
            ) AS tecnico_encargado
        FROM hojas_trabajo h
        JOIN proyectos pr ON h.id_proyecto = pr.id
        JOIN clientes cl ON pr.id_cliente = cl.id
        WHERE h.id = ?
    """, (id,), fetchone=True)

    if not hoja:
        flash("Hoja de trabajo no encontrada.", "danger")
        return redirect(url_for("hojas_trabajo_bp.listar_hojas"))
    
    if request.method == "POST":
        tecnico_encargado = request.form.get("tecnico_encargado")
        db.execute_query("""
            UPDATE hojas_trabajo SET
                tecnico_encargado = ?
            WHERE id = ?
        """, (tecnico_encargado, id))
        db.execute_query("DELETE FROM capitulos_hojas WHERE id_hoja = ?", (id,))
        db.execute_query("DELETE FROM partidas_hojas WHERE id_hoja = ?", (id,))

        form_dict = request.form.to_dict(flat=False)
        chapters = extract_chapters(form_dict, "descripcion")
        for chap_idx, chap_data in chapters.items():
            chapter_number = str(int(chap_idx) + 1)
            db.execute_query(
                "INSERT INTO capitulos_hojas (id_hoja, numero, descripcion) VALUES (?, ?, ?)",
                (id, chapter_number, chap_data["descripcion"])
            )
            for i, part_data in enumerate(chap_data["partidas"]):
                partida_number = f"{chapter_number}.{i+1}"
                cantidad = float(part_data.get("cantidad", 0) or 0)
                precio = float(part_data.get("precio", 0) or 0)
                total = float(part_data.get("total", 0) or 0)
                margen = float(part_data.get("margen", 0) or 0)
                final = float(part_data.get("final", 0) or 0)
                db.execute_query("""
                    INSERT INTO partidas_hojas 
                    (id_hoja, capitulo_numero, descripcion, unitario, cantidad, precio, total, margen, final)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    id,
                    partida_number,
                    part_data.get("descripcion", ""),
                    part_data.get("unitario", ""),
                    cantidad,
                    precio,
                    total,
                    margen,
                    final
                ))
        flash("Hoja de trabajo actualizada correctamente.", "success")
        return redirect(url_for("hojas_trabajo_bp.listar_hojas"))
    
    capitulos_rows = db.execute_query("SELECT * FROM capitulos_hojas WHERE id_hoja = ? ORDER BY id", (id,), fetch=True)
    capitulos = []
    for cap in capitulos_rows:
        cap_dict = dict(cap)
        cap_number = cap_dict.get("numero", "")
        partidas_rows = db.execute_query(
            "SELECT * FROM partidas_hojas WHERE id_hoja = ? AND capitulo_numero LIKE ? ORDER BY id",
            (id, f"{cap_number}.%"), fetch=True)
        cap_dict["partidas"] = [dict(p) for p in partidas_rows]
        capitulos.append(cap_dict)
    
    return render_template("hojas_trabajo_editar.html", hoja=hoja, capitulos=capitulos)

@hojas_trabajo_bp.route("/hojas_trabajo/borrar/<int:id>", methods=["POST"])
def borrar_hoja_trabajo(id):
    db.execute_query("DELETE FROM hojas_trabajo WHERE id = ?", (id,))
    flash("Hoja de trabajo borrada correctamente.", "success")
    return redirect(url_for("hojas_trabajo_bp.listar_hojas"))

@hojas_trabajo_bp.route("/hojas_trabajo/clonar/<int:id>", methods=["POST"])
def clonar_hoja_trabajo(id):
    original = db.execute_query("SELECT * FROM presupuestos WHERE id = ?", (id,), fetchone=True)
    if not original:
        flash("Presupuesto no encontrado.", "danger")
        return redirect(url_for("presupuestos_bp.listar_presupuestos"))
    
    base_ref = re.sub(r" TR\d+$", "", original["referencia"])
    clones = db.execute_query("SELECT referencia FROM hojas_trabajo WHERE referencia LIKE ?", (f"{base_ref} TR%",), fetch=True)
    if clones:
        versions = []
        for clone in clones:
            ref = clone["referencia"]
            m = re.search(r" TR(\d+)$", ref)
            if m:
                try:
                    versions.append(int(m.group(1)))
                except ValueError:
                    pass
        new_version = max(versions) + 1 if versions else 1
    else:
        new_version = 1
        
    new_ref = f"{base_ref} TR{new_version}"
    fecha = datetime.now().strftime("%Y-%m-%d")
    new_hoja_id = db.execute_query("""
        INSERT INTO hojas_trabajo (
            id_proyecto, referencia, fecha, tipo_via, nombre_via, numero_via, puerta,
            codigo_postal, poblacion, titulo, notas, tecnico_encargado, aprobacion, fecha_aprobacion, estado
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        original["id_proyecto"],
        new_ref,
        fecha,
        original["tipo_via"],
        original["nombre_via"],
        original["numero_via"],
        original["puerta"],
        original["codigo_postal"],
        original["poblacion"],
        original["titulo"],
        original["notas"],
        original["tecnico_encargado"],
        original["aprobacion"],
        original["fecha_aprobacion"],
        original["estado"]
    ))
    
    capitulos = db.execute_query("SELECT * FROM capitulos WHERE id_presupuesto = ? ORDER BY id", (id,), fetch=True)
    for cap in capitulos:
        db.execute_query("""
            INSERT INTO capitulos_hojas (id_hoja, numero, descripcion)
            VALUES (?, ?, ?)
        """, (new_hoja_id, cap["numero"], cap["descripcion"]))
        partidas = db.execute_query("SELECT * FROM partidas WHERE id_presupuesto = ? AND capitulo_numero LIKE ? ORDER BY id",
                                     (id, f"{cap['numero']}.%"), fetch=True)
        for part in partidas:
            db.execute_query("""
                INSERT INTO partidas_hojas (id_hoja, capitulo_numero, descripcion, unitario, cantidad, precio, total, margen, final)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                new_hoja_id,
                part["capitulo_numero"],
                part["descripcion"],
                part["unitario"],
                part["cantidad"],
                part["precio"],
                part["total"],
                part["margen"],
                part["final"]
            ))
    
    flash(f"Hoja de trabajo creada correctamente. Nueva referencia: {new_ref}", "success")
    return redirect(url_for("hojas_trabajo_bp.listar_hojas"))

@hojas_trabajo_bp.route("/hojas_trabajo/excel/<int:id>")
def exportar_excel_hoja_trabajo(id):
    hoja = db.execute_query("SELECT * FROM hojas_trabajo WHERE id = ?", (id,), fetchone=True)
    if not hoja:
        flash("Hoja de trabajo no encontrada.", "danger")
        return redirect(url_for("hojas_trabajo_bp.listar_hojas"))
    
    capitulos = db.execute_query("SELECT * FROM capitulos_hojas WHERE id_hoja = ? ORDER BY id", (id,), fetch=True) or []
    partidas = db.execute_query("SELECT * FROM partidas_hojas WHERE id_hoja = ? ORDER BY id", (id,), fetch=True) or []
    
    capitulos_dict = {}
    for cap in capitulos:
        capitulos_dict[cap["numero"]] = {"descripcion": cap["descripcion"], "partidas": []}
    for part in partidas:
        key = part["capitulo_numero"].split('.')[0]
        if key in capitulos_dict:
            capitulos_dict[key]["partidas"].append(dict(part))
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoja de Trabajo"
    
    ws["A1"] = "Referencia"
    ws["B1"] = hoja["referencia"]
    ws["A2"] = "Título"
    ws["B2"] = hoja["titulo"]
    ws["A3"] = "Fecha"
    ws["B3"] = hoja["fecha"]
    
    current_row = 5
    for cap_num in sorted(capitulos_dict.keys(), key=lambda x: float(x) if x.replace('.', '', 1).isdigit() else x):
        cap_data = capitulos_dict[cap_num]
        ws.cell(row=current_row, column=1, value=f"Capítulo {cap_num}")
        ws.cell(row=current_row, column=2, value=cap_data["descripcion"])
        current_row += 1
        if cap_data["partidas"]:
            headers = ["Descripción", "Unitario", "Cantidad", "Precio (€)", "Total (€)", "Margen (%)", "Final (€)"]
            for j, header in enumerate(headers, start=2):
                ws.cell(row=current_row, column=j, value=header)
            current_row += 1
            for part in cap_data["partidas"]:
                r = current_row
                ws.cell(row=r, column=2, value=part.get("descripcion"))
                ws.cell(row=r, column=3, value=part.get("unitario"))
                cantidad = float(part.get("cantidad") or 0)
                precio = float(part.get("precio") or 0)
                ws.cell(row=r, column=4, value=cantidad)
                ws.cell(row=r, column=5, value=precio)
                ws.cell(row=r, column=6, value=f"=D{r}*E{r}")
                margen = float(part.get("margen") or 40)
                ws.cell(row=r, column=7, value=margen)
                ws.cell(row=r, column=8, value=f"=F{r}*(1+G{r}/100)")
                current_row += 1
        else:
            ws.cell(row=current_row, column=2, value="Sin partidas")
            current_row += 1
        current_row += 1

    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = max_length + 2
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    file_name = f"HojaTrabajo_{hoja['referencia']}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=file_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@hojas_trabajo_bp.route("/hojas_trabajo/ver_pdf/<int:id>")
def ver_pdf_hoja_trabajo(id):
    hoja = db.execute_query("SELECT * FROM hojas_trabajo WHERE id = ?", (id,), fetchone=True)
    if not hoja:
        flash("Hoja de trabajo no encontrada.", "danger")
        return redirect(url_for("hojas_trabajo_bp.listar_hojas"))
    
    capitulos = db.execute_query("SELECT * FROM capitulos_hojas WHERE id_hoja = ? ORDER BY id", (id,), fetch=True) or []
    partidas = db.execute_query("SELECT * FROM partidas_hojas WHERE id_hoja = ? ORDER BY id", (id,), fetch=True) or []
    
    capitulos_dict = {}
    for cap in capitulos:
        capitulos_dict[cap["numero"]] = {"descripcion": cap["descripcion"], "partidas": []}
    for part in partidas:
        key = part["capitulo_numero"].split('.')[0]
        if key in capitulos_dict:
            capitulos_dict[key]["partidas"].append(dict(part))
    
    return render_template("hoja_trabajo_pdf.html", hoja=hoja, capitulos=capitulos_dict)
