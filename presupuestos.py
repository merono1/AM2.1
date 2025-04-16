from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, Response
from dbhelper import DBHelper, create_schema
from datetime import datetime
import re
import io
import os
import pdfkit  # Se mantiene para otros usos si es necesario
import openpyxl  # Para generar Excel
from openpyxl.utils import get_column_letter

# Importa WeasyPrint para generar PDF sin dependencias externas a nivel de ejecutable
from weasyprint import HTML

presupuestos_bp = Blueprint("presupuestos_bp", __name__)
db = DBHelper()

def generar_referencia_presupuesto(proyecto_id):
    fecha = datetime.now().strftime("%d%m%y")
    tipo = db.execute_query("SELECT tipo_proyecto FROM proyectos WHERE id = ?", (proyecto_id,), fetchone=True)
    if tipo:
        tipo_proyecto = tipo[0]
        letras = (tipo_proyecto[0] + tipo_proyecto[2] + tipo_proyecto[4]).upper() if len(tipo_proyecto) >= 5 else tipo_proyecto[:3].upper()
    else:
        letras = "XXX"
    base_ref = f"PR{int(proyecto_id):03d}{letras}-{fecha}-"
    similares = db.execute_query("SELECT referencia FROM presupuestos WHERE referencia LIKE ?", (f"{base_ref}%",), fetch=True)
    sufijo = f"{len(similares) + 1:02d}"
    return base_ref + sufijo

@presupuestos_bp.route("/presupuestos")
def listar_presupuestos():
    presupuestos = db.execute_query("""
        SELECT p.id, p.referencia, p.fecha, cl.nombre, pr.tipo_proyecto, pr.nombre_proyecto,
               p.tecnico_encargado, p.aprobacion, p.fecha_aprobacion, p.estado
        FROM presupuestos p
        JOIN proyectos pr ON p.id_proyecto = pr.id
        JOIN clientes cl ON pr.id_cliente = cl.id
        ORDER BY p.id DESC
    """, fetch=True)
    return render_template("presupuestos_list.html", presupuestos=presupuestos)

@presupuestos_bp.route("/presupuestos/nuevo", methods=["GET", "POST"])
def nuevo_presupuesto():
    proyectos = db.execute_query("SELECT id, nombre_proyecto FROM proyectos", fetch=True)
    proyecto_id_param = request.args.get("proyecto_id")
    if proyecto_id_param:
        try:
            proyecto_id_param = int(proyecto_id_param)
        except ValueError:
            proyecto_id_param = None
    datos_proyecto = {}
    if proyecto_id_param:
        proyecto_info = db.execute_query("""
            SELECT pr.nombre_proyecto, pr.calle, pr.nombre_via, pr.numero, pr.puerta,
                   pr.codigo_postal, pr.poblacion, cl.nombre
            FROM proyectos pr
            JOIN clientes cl ON pr.id_cliente = cl.id
            WHERE pr.id = ?
        """, (proyecto_id_param,), fetchone=True)
        if proyecto_info:
            datos_proyecto = {
                "titulo": proyecto_info[0],
                "tipo_via": proyecto_info[1],
                "nombre_via": proyecto_info[2],
                "numero_via": proyecto_info[3],
                "puerta": proyecto_info[4],
                "codigo_postal": proyecto_info[5],
                "poblacion": proyecto_info[6],
                "cliente_nombre": proyecto_info[7],
                "referencia": generar_referencia_presupuesto(proyecto_id_param)
            }
            datos_proyecto["proyecto_id"] = proyecto_id_param

    if request.method == "POST":
        proyecto_id = request.form.get("proyecto_id")
        if not proyecto_id:
            flash("Debe seleccionar un proyecto.", "danger")
            return render_template("presupuesto_nuevo.html", proyectos=proyectos,
                                   capitulos=[], proyecto_id_param=proyecto_id_param, **datos_proyecto)
        try:
            proyecto_id_int = int(proyecto_id)
        except ValueError:
            proyecto_id_int = None

        titulo = request.form.get("titulo")
        tipo_via = request.form.get("tipo_via")
        nombre_via = request.form.get("nombre_via")
        numero_via = request.form.get("numero_via")
        puerta = request.form.get("puerta")
        codigo_postal = request.form.get("codigo_postal")
        poblacion = request.form.get("poblacion")
        notas = request.form.get("notas")
        fecha = datetime.now().strftime("%Y-%m-%d")
        referencia = generar_referencia_presupuesto(proyecto_id_int)

        presupuesto_id = db.execute_query("""
            INSERT INTO presupuestos (
                id_proyecto, referencia, fecha, tipo_via, nombre_via, numero_via, puerta,
                codigo_postal, poblacion, titulo, notas
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (proyecto_id_int, referencia, fecha, tipo_via, nombre_via, numero_via, puerta,
              codigo_postal, poblacion, titulo, notas))

        form_dict = request.form.to_dict(flat=False)
        chapters = {}
        chapter_pattern = re.compile(r'capitulos\[(\d+)\]\[tipo\]')
        for key, values in form_dict.items():
            m = chapter_pattern.match(key)
            if m:
                cap_idx = m.group(1)
                chapters[cap_idx] = {'tipo': values[0], 'partidas': []}
        partida_pattern = re.compile(r'capitulos\[(\d+)\]\[partidas\]\[(\d+)\]\[([^\]]+)\]')
        partidas_group = {}
        for key, values in form_dict.items():
            m = partida_pattern.match(key)
            if m:
                cap_idx = m.group(1)
                part_idx = m.group(2)
                field = m.group(3)
                if cap_idx not in partidas_group:
                    partidas_group[cap_idx] = {}
                if part_idx not in partidas_group[cap_idx]:
                    partidas_group[cap_idx][part_idx] = {}
                partidas_group[cap_idx][part_idx][field] = values[0]
        for cap_idx, parts in partidas_group.items():
            if cap_idx in chapters:
                for part_idx in sorted(parts.keys(), key=int):
                    chapters[cap_idx]['partidas'].append(parts[part_idx])
        for cap_idx, chapter in chapters.items():
            chapter_number = str(int(cap_idx) + 1)
            db.execute_query("""
                INSERT INTO capitulos (id_presupuesto, numero, descripcion)
                VALUES (?, ?, ?)
            """, (presupuesto_id, chapter_number, chapter['tipo']))
            for i, partida in enumerate(chapter['partidas']):
                cap_num = f"{chapter_number}.{i+1}"
                cantidad = float(partida.get('cantidad') or 0)
                precio = float(partida.get('precio') or 0)
                total = float(partida.get('total') or 0)
                margen = float(partida.get('margen') or 40)
                final = total * (1 + margen / 100)
                db.execute_query("""
                    INSERT INTO partidas (id_presupuesto, capitulo_numero, descripcion, unitario, cantidad, precio, total, margen, final)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (presupuesto_id, cap_num, partida.get('descripcion'), partida.get('unitario'),
                      cantidad, precio, total, margen, final))
        flash(f"Presupuesto creado correctamente. Referencia: {referencia}", "success")
        return redirect(url_for("presupuestos_bp.listar_presupuestos"))

    return render_template("presupuesto_nuevo.html", proyectos=proyectos,
                           proyecto_id_param=proyecto_id_param, **datos_proyecto, capitulos=[])

@presupuestos_bp.route("/presupuestos/editar/<int:id>", methods=["GET", "POST"])
def editar_presupuesto(id):
    presupuesto = db.execute_query("SELECT * FROM presupuestos WHERE id = ?", (id,), fetchone=True)
    if request.method == "POST":
        titulo = request.form.get("titulo")
        tipo_via = request.form.get("tipo_via")
        nombre_via = request.form.get("nombre_via")
        numero_via = request.form.get("numero_via")
        puerta = request.form.get("puerta")
        codigo_postal = request.form.get("codigo_postal")
        poblacion = request.form.get("poblacion")
        notas = request.form.get("notas")
        db.execute_query("""
            UPDATE presupuestos SET
                titulo = ?, tipo_via = ?, nombre_via = ?, numero_via = ?, puerta = ?,
                codigo_postal = ?, poblacion = ?, notas = ?
            WHERE id = ?
        """, (titulo, tipo_via, nombre_via, numero_via, puerta, codigo_postal, poblacion, notas, id))
        db.execute_query("DELETE FROM capitulos WHERE id_presupuesto = ?", (id,))
        db.execute_query("DELETE FROM partidas WHERE id_presupuesto = ?", (id,))
        form_dict = request.form.to_dict(flat=False)
        chapters = {}
        chapter_pattern = re.compile(r'capitulos\[(\d+)\]\[tipo\]')
        for key, values in form_dict.items():
            m = chapter_pattern.match(key)
            if m:
                cap_idx = m.group(1)
                chapters[cap_idx] = {'tipo': values[0], 'partidas': []}
        partida_pattern = re.compile(r'capitulos\[(\d+)\]\[partidas\]\[(\d+)\]\[([^\]]+)\]')
        partidas_group = {}
        for key, values in form_dict.items():
            m = partida_pattern.match(key)
            if m:
                cap_idx = m.group(1)
                part_idx = m.group(2)
                field = m.group(3)
                if cap_idx not in partidas_group:
                    partidas_group[cap_idx] = {}
                if part_idx not in partidas_group[cap_idx]:
                    partidas_group[cap_idx][part_idx] = {}
                partidas_group[cap_idx][part_idx][field] = values[0]
        for cap_idx, parts in partidas_group.items():
            if cap_idx in chapters:
                for part_idx in sorted(parts.keys(), key=int):
                    chapters[cap_idx]['partidas'].append(parts[part_idx])
        for cap_idx, chapter in chapters.items():
            chapter_number = str(int(cap_idx) + 1)
            db.execute_query("""
                INSERT INTO capitulos (id_presupuesto, numero, descripcion)
                VALUES (?, ?, ?)
            """, (id, chapter_number, chapter['tipo']))
            for i, partida in enumerate(chapter['partidas']):
                cap_num = f"{chapter_number}.{i+1}"
                cantidad = float(partida.get('cantidad') or 0)
                precio = float(partida.get('precio') or 0)
                total = float(partida.get('total') or 0)
                margen = float(partida.get('margen') or 40)
                final = total * (1 + margen / 100)
                db.execute_query("""
                    INSERT INTO partidas (id_presupuesto, capitulo_numero, descripcion, unitario, cantidad, precio, total, margen, final)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (id, cap_num, partida.get('descripcion'), partida.get('unitario'),
                      cantidad, precio, total, margen, final))
        flash("Presupuesto actualizado correctamente.", "success")
        return redirect(url_for("presupuestos_bp.listar_presupuestos"))
    capitulos_rows = db.execute_query("SELECT * FROM capitulos WHERE id_presupuesto = ? ORDER BY id", (id,), fetch=True) or []
    capitulos = []
    for row in capitulos_rows:
        cap = dict(row)
        cap["tipo"] = cap["descripcion"]
        partidas_rows = db.execute_query(
            "SELECT * FROM partidas WHERE id_presupuesto = ? AND capitulo_numero LIKE ? ORDER BY id",
            (id, f"{cap['numero']}.%"), fetch=True
        ) or []
        cap["partidas"] = [dict(p) for p in partidas_rows]
        capitulos.append(cap)
    proyecto = db.execute_query("SELECT nombre_proyecto FROM proyectos WHERE id = ?", (presupuesto["id_proyecto"],), fetchone=True)
    proyecto_nombre = proyecto[0] if proyecto else ""
    return render_template("presupuesto_nuevo.html", presupuesto=presupuesto,
                           capitulos=capitulos, proyecto_id=presupuesto["id_proyecto"],
                           proyecto_nombre=proyecto_nombre)

@presupuestos_bp.route("/presupuestos/actualizar/<int:id>", methods=["POST"])
def actualizar_presupuesto_estado(id):
    tecnico = request.form.get("tecnico")
    aprobacion = request.form.get("aprobacion")
    fecha_aprobacion = request.form.get("fecha_aprobacion")
    estado = request.form.get("estado")
    db.execute_query("""
        UPDATE presupuestos SET 
            tecnico_encargado = ?, 
            aprobacion = ?, 
            fecha_aprobacion = ?, 
            estado = ?
        WHERE id = ?
    """, (tecnico, aprobacion, fecha_aprobacion, estado, id))
    return "OK"

@presupuestos_bp.route("/presupuestos/borrar/<int:id>", methods=["POST"])
def borrar_presupuesto(id):
    db.execute_query("DELETE FROM presupuestos WHERE id = ?", (id,))
    flash("Presupuesto borrado correctamente.", "success")
    return redirect(url_for("presupuestos_bp.listar_presupuestos"))

@presupuestos_bp.route("/presupuestos/clonar/<int:id>", methods=["POST"])
def clonar_presupuesto(id):
    original = db.execute_query("SELECT * FROM presupuestos WHERE id = ?", (id,), fetchone=True)
    if not original:
        flash("Presupuesto no encontrado.", "danger")
        return redirect(url_for("presupuestos_bp.listar_presupuestos"))
    base_ref = re.sub(r" V\d+$", "", original["referencia"])
    clones = db.execute_query("SELECT referencia FROM presupuestos WHERE referencia LIKE ?", (f"{base_ref} V%",), fetch=True)
    if clones:
        versions = []
        for clone in clones:
            ref = clone["referencia"]
            m = re.search(r" V(\d+)$", ref)
            if m:
                try:
                    ver = int(m.group(1))
                    versions.append(ver)
                except ValueError:
                    pass
        new_version = max(versions) + 1 if versions else 2
    else:
        new_version = 2
    new_ref = f"{base_ref} V{new_version}"
    fecha = datetime.now().strftime("%Y-%m-%d")
    new_presupuesto_id = db.execute_query("""
        INSERT INTO presupuestos (
            id_proyecto, referencia, fecha, tipo_via, nombre_via, numero_via, puerta,
            codigo_postal, poblacion, titulo, notas
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (original["id_proyecto"], new_ref, fecha, original["tipo_via"],
          original["nombre_via"], original["numero_via"], original["puerta"],
          original["codigo_postal"], original["poblacion"], original["titulo"], original["notas"]))
    capitulos = db.execute_query("SELECT * FROM capitulos WHERE id_presupuesto = ? ORDER BY id", (id,), fetch=True)
    for cap in capitulos:
        db.execute_query("""
            INSERT INTO capitulos (id_presupuesto, numero, descripcion)
            VALUES (?, ?, ?)
        """, (new_presupuesto_id, cap["numero"], cap["descripcion"]))
        partidas = db.execute_query("SELECT * FROM partidas WHERE id_presupuesto = ? AND capitulo_numero LIKE ? ORDER BY id",
                                     (id, f"{cap['numero']}.%"), fetch=True)
        for part in partidas:
            db.execute_query("""
                INSERT INTO partidas (id_presupuesto, capitulo_numero, descripcion, unitario, cantidad, precio, total, margen, final)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (new_presupuesto_id, part["capitulo_numero"], part["descripcion"], part["unitario"],
                  part["cantidad"], part["precio"], part["total"], part["margen"], part["final"]))
    flash(f"Presupuesto clonado correctamente. Nueva referencia: {new_ref}", "success")
    return redirect(url_for("presupuestos_bp.listar_presupuestos"))

@presupuestos_bp.route("/presupuestos/por_proyecto/<int:proyecto_id>")
def listar_presupuestos_por_proyecto(proyecto_id):
    presupuestos = db.execute_query("""
        SELECT id, referencia, fecha, titulo, estado 
        FROM presupuestos
        WHERE id_proyecto = ?
        ORDER BY fecha DESC
    """, (proyecto_id,), fetch=True)
    return render_template("presupuestos_por_proyecto.html", presupuestos=presupuestos)

# -------------------- ENDPOINT PARA EXPORTAR A EXCEL --------------------
@presupuestos_bp.route("/presupuestos/excel/<int:id>", endpoint="exportar_excel_presupuesto")
def exportar_excel_presupuesto(id):
    presupuesto = db.execute_query("SELECT * FROM presupuestos WHERE id = ?", (id,), fetchone=True)
    if not presupuesto:
        flash("Presupuesto no encontrado.", "danger")
        return redirect(url_for("presupuestos_bp.listar_presupuestos"))
    capitulos = db.execute_query("SELECT * FROM capitulos WHERE id_presupuesto = ? ORDER BY id", (id,), fetch=True) or []
    partidas = db.execute_query("SELECT * FROM partidas WHERE id_presupuesto = ? ORDER BY id", (id,), fetch=True) or []
    capitulos_dict = {}
    for cap in capitulos:
        capitulos_dict[cap["numero"]] = {"descripcion": cap["descripcion"], "partidas": []}
    for part in partidas:
        part_dict = dict(part)
        chapter_key = part_dict["capitulo_numero"].split('.')[0]
        if chapter_key in capitulos_dict:
            capitulos_dict[chapter_key]["partidas"].append(part_dict)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Presupuesto"
    ws["A1"] = "Referencia"
    ws["B1"] = presupuesto["referencia"]
    ws["A2"] = "Título"
    ws["B2"] = presupuesto["titulo"]
    ws["A3"] = "Fecha"
    ws["B3"] = presupuesto["fecha"]
    current_row = 5
    for cap_num in sorted(capitulos_dict.keys(), key=lambda x: float(x) if x.replace('.', '', 1).isdigit() else x):
        cap_data = capitulos_dict[cap_num]
        ws.cell(row=current_row, column=1, value=f"Capítulo {cap_num}")
        ws.cell(row=current_row, column=2, value=cap_data["descripcion"])
        current_row += 1
        if cap_data["partidas"]:
            headers = ["Descripción", "Unitario", "Cantidad", "Precio (€)", "Total (€)", "Margen (%)", "Final (€)"]
            for j, header in enumerate(headers, start=2):
                ws.cell(row=current_row, column=j, value=header)
            current_row += 1
            for part in cap_data["partidas"]:
                ws.cell(row=current_row, column=2, value=part.get("descripcion"))
                ws.cell(row=current_row, column=3, value=part.get("unitario"))
                ws.cell(row=current_row, column=4, value=float(part.get("cantidad") or 0))
                ws.cell(row=current_row, column=5, value=float(part.get("precio") or 0))
                ws.cell(row=current_row, column=6, value=float(part.get("total") or 0))
                ws.cell(row=current_row, column=7, value=float(part.get("margen") or 40))
                try:
                    total_val = float(part.get("total") or 0)
                    margen_val = float(part.get("margen") or 40)
                    final_val = total_val * (1 + margen_val / 100)
                except Exception:
                    final_val = None
                ws.cell(row=current_row, column=8, value=final_val)
                current_row += 1
        else:
            ws.cell(row=current_row, column=2, value="Sin partidas")
            current_row += 1
        current_row += 1
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = max_length + 2
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"Presupuesto_{presupuesto['referencia']}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
# -------------------- FIN ENDPOINT EXCEL --------------------

# -------------------- NUEVO ENDPOINT PARA GENERAR PDF --------------------
@presupuestos_bp.route("/presupuestos/generar_pdf/<int:id>")
def generar_pdf_presupuesto(id):
    presupuesto = db.execute_query("SELECT * FROM presupuestos WHERE id = ?", (id,), fetchone=True)
    if not presupuesto:
        flash("Presupuesto no encontrado.", "danger")
        return redirect(url_for("presupuestos_bp.listar_presupuestos"))
    capitulos = db.execute_query("SELECT * FROM capitulos WHERE id_presupuesto = ? ORDER BY id", (id,), fetch=True) or []
    partidas = db.execute_query("SELECT * FROM partidas WHERE id_presupuesto = ? ORDER BY id", (id,), fetch=True) or []
    capitulos_dict = {}
    for cap in capitulos:
        capitulos_dict[cap["numero"]] = {"descripcion": cap["descripcion"], "partidas": []}
    for part in partidas:
        part_dict = dict(part)
        chapter_key = part_dict["capitulo_numero"].split('.')[0]
        if chapter_key in capitulos_dict:
            capitulos_dict[chapter_key]["partidas"].append(part_dict)
    rendered = render_template("presupuesto_pdf_lite.html", presupuesto=presupuesto, capitulos=capitulos_dict)
    # Usar WeasyPrint para generar el PDF sin dependencias de ejecutables externos
    pdf_data = HTML(string=rendered, base_url=request.base_url).write_pdf()
    if not pdf_data:
        flash("Error al generar el PDF.", "danger")
        return redirect(url_for("presupuestos_bp.listar_presupuestos"))
    output_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), "pdfs")
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    pdf_filename = os.path.join(output_folder, f"Presupuesto_{presupuesto['referencia']}.pdf")
    with open(pdf_filename, "wb") as f:
        f.write(pdf_data)
    # Configurar los encabezados para forzar la descarga
    response = Response(pdf_data, mimetype="application/octet-stream")
    response.headers["Content-Disposition"] = f"attachment; filename=Presupuesto_{presupuesto['referencia']}.pdf"
    return response
# -------------------- FIN NUEVO ENDPOINT PDF --------------------

@presupuestos_bp.route("/presupuestos/ver_pdf/<int:id>")
def ver_pdf_presupuesto(id):
    presupuesto = db.execute_query("SELECT * FROM presupuestos WHERE id = ?", (id,), fetchone=True)
    if not presupuesto:
        flash("Presupuesto no encontrado.", "danger")
        return redirect(url_for("presupuestos_bp.listar_presupuestos"))
    capitulos = db.execute_query("SELECT * FROM capitulos WHERE id_presupuesto = ? ORDER BY id", (id,), fetch=True) or []
    partidas = db.execute_query("SELECT * FROM partidas WHERE id = ? ORDER BY id", (id,), fetch=True) or []
    capitulos_dict = {}
    for cap in capitulos:
        capitulos_dict[cap["numero"]] = {"descripcion": cap["descripcion"], "partidas": []}
    for part in partidas:
        part_dict = dict(part)
        chapter_key = part_dict["capitulo_numero"].split('.')[0]
        if chapter_key in capitulos_dict:
            capitulos_dict[chapter_key]["partidas"].append(part_dict)
    return render_template("presupuesto_pdf_lite.html", presupuesto=presupuesto, capitulos=capitulos_dict)
